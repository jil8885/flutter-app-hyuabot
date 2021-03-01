import sqlite3
import openpyxl

path = "assets/databases/information.db"
xls_path = "database.xlsx"
connection = sqlite3.connect(path)
cursor = connection.cursor()
inschool = []
outschool = []

# 전화번호부(교내)
cursor.execute("select * from inschool")
for name, phone in cursor.fetchall():
    inschool.append((name, phone))

# 전화번호부(교외)
cursor.execute("select * from outschool")
for name, category, phone, latitude, longitude, menu in cursor.fetchall():
    outschool.append((name, category, phone, latitude, longitude, menu))

workbook = openpyxl.Workbook(write_only=True)

# 교내
worksheet = workbook.create_sheet("교내 데이터베이스")
worksheet.append(["name_ko", "name_en", "name_zh", "phone"])
for row in inschool:
    worksheet.append(row)

# 교외
worksheet = workbook.create_sheet("교외 데이터베이스")
worksheet.append(["name_ko", "name_en", "name_zh", "category_ko", "category_en", "category_zh", "phone", "latitude", "longitude", "menu_ko", "menu_en", "menu_zh"])
for row in outschool:
    name, category, phone, latitude, longitude, menu = row
    worksheet.append([name, None, None, category, None, None, phone, latitude, longitude, menu, None, None])

workbook.save(xls_path)
cursor.close()
connection.close()
